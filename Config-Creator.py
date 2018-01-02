#!/usr/bin/env python
'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
Install GIT (most likely you do not have the XLHELPER module. This has been forked to make it work with Python 2.7. It will ask you to download automatically)
Install OpenPYXL
Install FileInput

---VERSION---
VERSION 1.1
Currently Implemented Features
-Text based variable replacement
-XLSX based variable replacement
-Automatic math for CIDR on XLSX
	- If the XLSX column data contains a subnet in cidr format, and the variable contains IPADD, it will automatically switch it out for the gateway IP in subnet format
		- e.g. 10.1.1.0/24 -> 10.1.1.1 255.255.255.0
		- Additional variables would be HSRPPRI (changes to 10.1.1.2) and HSRPSEC (changes to 10.1.1.3)
	- Full list of math based variables (IPADD,HSRPPRI,HSRPSEC)


Features planned in the near future


'''

'''IMPORT MODULES'''
import os
import sys
#
try:
	from openpyxl import load_workbook
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
#	
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper
#
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()

# Netmask Function
from socket import inet_ntoa
from struct import pack

def calcDottedNetmask(mask):
    bits = 0xffffffff ^ (1 << 32 - mask) - 1
    return inet_ntoa(pack('>I', bits))


	
	


# Import Excel File Method
XLSimportq = raw_input('Import XLS file for data entry (Y/N)? ')
if 'y' in XLSimportq.lower():
	excelfilelocation = raw_input('File to load the excel data from (e.g. C:/Python27/exceldata.xlsx):')
	if excelfilelocation == '':
		excelfilelocation = 'C:/Python27/exceldata.xlsx'
	excelfilelocation = excelfilelocation.replace('"', '')

# Import TXT/CLI Method
if 'n' in XLSimportq.lower():
	templatefilelocation = raw_input('File to load the default template from (e.g. C:/Python27/basetemplate.txt):')
	if templatefilelocation == '':
		templatefilelocation = 'C:/Python27/basetemplate.txt'
	templatefilelocation = templatefilelocation.replace('"', '')
	try:
		with open(templatefilelocation, 'r') as file :
			configfiledata = file.read()
	except:
		print 'Error in loading the base template. Please enter a good path for the file.'
	# Import Variable List
	variablefilelocation = raw_input("File to load variables from (e.g. C:/Python27/variables.txt):")
	if variablefilelocation == '':
		variablefilelocation = 'C:/Python27/variables.txt'
	variablefilelocation = variablefilelocation.replace('"', '')
	try:
		with open(variablefilelocation) as f:
			variablelist = f.read().splitlines() 
	except:
		print 'Error in loading the variable list file. Please enter a good path for the file.'
	finalconfigpath = 'C:/Python27/' + excelmatchq2 + '.txt'
	finalconfigfile = raw_input('Configuration file path to save to (e.g. ' + finalconfigpath + '):')
	if finalconfigfile == '':
		finalconfigfile = finalconfigpath
	finalconfigfile = finalconfigfile.replace('"', '')

# Save File


if 'n' in XLSimportq.lower():
# Find and Replace
	for variable in variablelist:
		if not '#NOTE:' in variable:
			variablename = variable.rsplit(':',1)[1]
			variablename = variablename.strip()
			variableq = 'Please enter the value for the following variable ' + variablename + ' :'
			variablea = raw_input(variableq)
			configfiledata = configfiledata.replace(variablename,variablea)
if 'y' in XLSimportq.lower():
	for hosts in xlhelper.sheet_to_dict(excelfilelocation,'Name-Template'):
		OutputPath = hosts.get('Config Output').encode('utf-8')
		Hostname = hosts.get('Hostname').encode('utf-8')
		TemplatePath = hosts.get('Template Location').encode('utf-8')
		OutputSave = OutputPath + '\\' + Hostname + '.txt'
		try:
			with open(TemplatePath, 'r') as file :
				configfiledata = file.read()
		except:
			print 'Error in loading the base template. Please enter a good path for the file.'
		for variables in xlhelper.sheet_to_dict(excelfilelocation,'Variables'):
			xlscolumnname = variables.get('Variable Name').encode('utf-8')
			#
			for values in xlhelper.sheet_to_dict(excelfilelocation,'Values'):
				matchvalue = values.get('#HOSTNAME#').encode('utf-8')
				if matchvalue == Hostname:
					try:
						variableresult = values.get(xlscolumnname).encode('utf-8')
					except:
						variableresult = values.get(xlscolumnname)
						variableresult = str(variableresult)
					if "/" in variableresult and "IPADD" in xlscolumnname:
						variableresult = variableresult.split('/')
						cidrvariable = variableresult[1]
						cidrvariable = int(cidrvariable)
						subnetmask = calcDottedNetmask(cidrvariable)
						networkaddress = variableresult[0]
						networkaddress = networkaddress.split('.')
						firstoctet = networkaddress[0]
						secondoctet = networkaddress[1]
						thirdoctet = networkaddress[2]
						lastoctet = networkaddress[3]
						lastoctet = int(lastoctet)
						lastoctet = lastoctet + 1
						lastoctet = str(lastoctet)
						gatewayip = firstoctet + '.' + secondoctet + '.' + thirdoctet + '.' + lastoctet
						variableresult = str(gatewayip) + ' ' + str(subnetmask)
					if "/" in variableresult and "HSRPPRI" in xlscolumnname:
						variableresult = variableresult.split('/')
						cidrvariable = variableresult[1]
						cidrvariable = int(cidrvariable)
						subnetmask = calcDottedNetmask(cidrvariable)
						networkaddress = variableresult[0]
						networkaddress = networkaddress.split('.')
						firstoctet = networkaddress[0]
						secondoctet = networkaddress[1]
						thirdoctet = networkaddress[2]
						lastoctet = networkaddress[3]
						lastoctet = int(lastoctet)
						lastoctet = lastoctet + 2
						lastoctet = str(lastoctet)
						gatewayip = firstoctet + '.' + secondoctet + '.' + thirdoctet + '.' + lastoctet
						variableresult = str(gatewayip) + ' ' + str(subnetmask)
					if "/" in variableresult and "HSRPSEC" in xlscolumnname:
						variableresult = variableresult.split('/')
						cidrvariable = variableresult[1]
						cidrvariable = int(cidrvariable)
						subnetmask = calcDottedNetmask(cidrvariable)
						networkaddress = variableresult[0]
						networkaddress = networkaddress.split('.')
						firstoctet = networkaddress[0]
						secondoctet = networkaddress[1]
						thirdoctet = networkaddress[2]
						lastoctet = networkaddress[3]
						lastoctet = int(lastoctet)
						lastoctet = lastoctet + 3
						lastoctet = str(lastoctet)
						gatewayip = firstoctet + '.' + secondoctet + '.' + thirdoctet + '.' + lastoctet
						variableresult = str(gatewayip) + ' ' + str(subnetmask)
					variablereplace = variableresult
			configfiledata = configfiledata.replace(xlscolumnname,variablereplace)
			variablereplace = ''
			xlscolumnname = ''
			setname = ''
		print 'Writing config file for ' + Hostname
		with open(OutputSave, 'w') as file:
			file.write(configfiledata)
		print 'Done...'	
if 'n' in XLSimportq.lower():
	print 'Script nearing completion. Saving final configuration file.'
	with open(finalconfigfile, 'w') as file:
		file.write(configfiledata)
	print 'Done...'
# Exit